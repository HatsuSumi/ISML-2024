import json
import csv
import pandas as pd
from typing import Dict
from validate_data import CharacterDataValidator
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Font, Alignment

def load_cv_data(cv_file: str) -> Dict[str, str]:
    cv_data = {}
    try:
        with open(cv_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get('角色', '').strip()
                cv = row.get('CV', '').strip()
                if name:
                    cv_data[name] = cv
    except Exception as e:
        print(f"\n❌ Error loading CV data: {str(e)}")
        return {}
    return cv_data

def load_character_data(excel_file: str) -> pd.DataFrame:
    df = pd.read_excel(excel_file, dtype={'ID': str})
    
    column_map = {
        '角色': 'name',
        '作品': 'series',
        '头像': 'image',
        '状态': 'status',
        '分组': 'group',
        '季节': 'season',
        '性别': 'gender',
        'ID': 'id'
    }
    
    df_processed = df.rename(columns=column_map)
    return df_processed

def generate_character_data() -> Dict:
    data = {
        'stellar': {
            'female': [],
            'male': []
        },
        'nova': {
            season: {'female': [], 'male': []} 
            for season in ['winter', 'spring', 'summer', 'autumn']
        }
    }
    
    df = load_character_data('data/ISML2024_characters_updated.xlsx')
    
    for _, row in df.iterrows():
        name = row['name'].strip()
        character = {
            'id': str(row['id']),
            'name': name,
            'series': row['series'].strip(),
            'image': row['image'] if pd.notna(row['image']) else '',
            'status': row['status'] if pd.notna(row['status']) else '',
            'cv': row['CV'] if pd.notna(row['CV']) else ''  
        }
        
        if row['group'] == 'stellar':
            data['stellar'][row['gender']].append(character)
        else:  # nova
            season = row['season'].lower()
            gender = row['gender'].lower()
            if season in data['nova'] and gender in data['nova'][season]:
                data['nova'][season][gender].append(character)

    for gender in ['female', 'male']:
        data['stellar'][gender].sort(key=lambda x: x['id'])
    
    for season in data['nova']:
        for gender in ['female', 'male']:
            data['nova'][season][gender].sort(key=lambda x: x['id'])
    
    return data

def print_statistics(data: Dict) -> None:
    stellar_f = len(data['stellar']['female'])
    stellar_m = len(data['stellar']['male'])
    
    nova_counts = {
        season: {
            'female': len(data['nova'][season]['female']),
            'male': len(data['nova'][season]['male'])
        }
        for season in data['nova']
    }
    
    total_nova = sum(
        count['female'] + count['male']
        for count in nova_counts.values()
    )

    print("\n📊 Character Statistics:")
    print(f"✨ Stellar Division:")
    print(f"   Female: {stellar_f}")
    print(f"   Male: {stellar_m}")
    print(f"\n🌟 Nova Division:")
    for season in nova_counts:
        print(f"   {season.title()}:")
        print(f"      Female: {nova_counts[season]['female']}")
        print(f"      Male: {nova_counts[season]['male']}")
    
    total = stellar_f + stellar_m + total_nova
    print(f"\n📈 Total Characters: {total}")

def load_image_data(img_file: str) -> Dict[str, str]:
    img_data = {}
    try:
        with open(img_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get('角色', '').strip()
                img = row.get('头像', '').strip()
                if name:
                    img_data[name] = img
    except Exception as e:
        print(f"\n❌ Error loading image data: {str(e)}")
        return {}
    return img_data

def update_excel_data(excel_file: str, cv_data: Dict[str, str], img_data: Dict[str, str]) -> None:
    output_excel = 'data/ISML2024_characters_updated.xlsx'
    
    try:
        original_excel = pd.ExcelFile(excel_file)
        df = pd.read_excel(original_excel, dtype={'ID': str})
        
        original_wb = load_workbook(excel_file)
        original_sheet = original_wb.active
        
        # 更新数据但保持原始格式
        df['CV'] = df['角色'].map(lambda x: cv_data.get(x.strip(), ''))
        df['头像'] = df['角色'].map(lambda x: img_data.get(x.strip(), ''))
        
        try:
            with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=original_excel.sheet_names[0], index=False)
                
                worksheet = writer.sheets[original_excel.sheet_names[0]]
                
                for col in original_sheet.column_dimensions:
                    original_width = original_sheet.column_dimensions[col].width
                    if original_width:
                        worksheet.column_dimensions[col].width = round(original_width, 2)
                
                for row in range(1, len(df) + 2):
                    worksheet.row_dimensions[row].height = 20
                
                cv_col = None
                img_col = None
                for idx, col_name in enumerate(df.columns, 1):
                    if col_name == 'CV':
                        cv_col = idx
                    elif col_name == '头像':
                        img_col = idx
                
                for row in range(1, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        source_cell = original_sheet.cell(row=row, column=col)
                        target_cell = worksheet.cell(row=row, column=col)
                        
                        target_cell.font = Font(
                            name='楷体',
                            size=12,
                            bold=source_cell.font.bold if source_cell.font else False,
                            italic=source_cell.font.italic if source_cell.font else False
                        )
                        
                        if source_cell.border:
                            target_cell.border = copy(source_cell.border)
                        
                        if col in (cv_col, img_col):
                            target_cell.alignment = Alignment(
                                horizontal='center',
                                vertical='center'
                            )
                        elif source_cell.alignment:
                            target_cell.alignment = Alignment(
                                horizontal=source_cell.alignment.horizontal,
                                vertical=source_cell.alignment.vertical,
                                wrap_text=source_cell.alignment.wrap_text
                            )
                            
        except PermissionError:
            print(f"\n❌ 无法保存文件 {output_excel}")
            print("请确保文件未被其他程序打开，然后重试")
            raise
            
    except Exception as e:
        print(f"\n❌ 处理Excel文件时出错: {str(e)}")
        raise

def main():
    try:
        cv_data = load_cv_data('data/ISML2024_characters_cv.csv')
        print("✅ CV data loaded successfully")
        
        img_data = load_image_data('data/common/ISML2024_characters_img.csv')
        print("✅ Image data loaded successfully")
        
        update_excel_data('data/common/ISML2024_characters.xlsx', cv_data, img_data)
        
        data = generate_character_data()
        print("✅ Character data generated")
        
        print_statistics(data)
        
        output_file = 'data/common/ISML2024_characters.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"\n✅ Data saved to: {output_file}")
        
        validator = CharacterDataValidator(output_file)
        
        if validator.validate():
            print(f"\n✅ Data validation passed")
        else:
            print(f"\n❌ Data validation failed")
            
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        raise

if __name__ == '__main__':
    main() 
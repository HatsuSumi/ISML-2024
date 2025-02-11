import csv
import pandas as pd
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def create_backup(filename):
    """创建文件备份"""
    source = Path(filename)
    if not source.exists():
        return
    
    # 创建备份文件夹
    backup_dir = Path('data/backups')
    backup_dir.mkdir(exist_ok=True)
    
    # 生成备份文件名（添加时间戳）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"{source.stem}_{timestamp}{source.suffix}"
    backup_path = backup_dir / backup_name
    
    # 复制文件
    shutil.copy2(source, backup_path)
    print(f"✅ 已创建备份: {backup_path}")

def load_additional_data(filename):
    data = {}
    try:
        with open(filename, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get('角色', '').strip()
                value = list(row.values())[1].strip() if len(row.values()) > 1 else ''
                if name:
                    data[name] = value
    except Exception as e:
        print(f"❌ 读取文件 {filename} 时出错: {str(e)}")
        return {}
    return data

def process_ranking_data(excel_file, cv_data, en_data, img_data):
    # 读取Excel文件
    df = pd.read_excel(excel_file)
    
    def clean_str(value):
        """清理字符串中的空格"""
        if pd.isna(value):
            return None
        return str(value).strip()
    
    characters = []
    for index, row in df.iterrows():
        name = clean_str(row['角色'])
        if not name:  # 如果角色名为空，跳过该行
            continue
        
        # 处理得票数
        votes = row['得票数']
        if pd.isna(votes) or str(votes).strip() == '-':
            votes = 0
        else:
            try:
                votes = int(votes)
            except (ValueError, TypeError):
                votes = 0
        
        characters.append({
            'rank': index + 1,
            'name': name,
            'series': clean_str(row['作品']),
            'votes': votes,
            'cv': cv_data.get(name, None),
            'name_en': en_data.get(name, None),
            'image': img_data.get(name, None)
        })
    
    # 按得票数排序
    characters.sort(key=lambda x: x['votes'], reverse=True)
    # 更新排名
    for i, char in enumerate(characters, 1):
        char['rank'] = i
        
    return characters

def save_to_excel(characters, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "角色排名"
    
    headers = ['排名', '角色', '作品', '得票数', 'CV', '角色（英）', '头像']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(name='微软雅黑', bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for row, char in enumerate(characters, 2):
        ws.cell(row=row, column=1, value=char['rank'])
        ws.cell(row=row, column=2, value=char['name'])
        ws.cell(row=row, column=3, value=char['series'])
        ws.cell(row=row, column=4, value=char['votes'])
        ws.cell(row=row, column=5, value=char['cv'])
        ws.cell(row=row, column=6, value=char['name_en'])
        ws.cell(row=row, column=7, value=char['image'])
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)

def main():
    # 创建备份
    create_backup('data/01-female-nomination.xlsx')
    
    # 加载补充数据
    cv_data = load_additional_data('data/ISML2024_characters_cv.csv')
    en_data = load_additional_data('data/ISML2024_characters_en.csv')
    img_data = load_additional_data('data/ISML2024_characters_img.csv')
    
    print("✅ 已加载补充数据")
    
    # 处理Excel数据并保存
    characters = process_ranking_data('data/01-female-nomination.xlsx', cv_data, en_data, img_data)
    save_to_excel(characters, 'data/ranking_results.xlsx')
    
    print("✅ 处理完成！Excel文件已保存到 data/ranking_results.xlsx")

if __name__ == '__main__':
    main() 